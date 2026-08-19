import hashlib
import io
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime, time
from typing import Iterator, Optional

import msoffcrypto
import openpyxl
import pyzipper

log = logging.getLogger("parser")

SHEET_HINTS = ["가계부 내역", "가계부내역", "가계부"]

COLUMN_ALIASES = {
    "date": ["날짜", "거래일자", "일자", "거래일"],
    "time": ["시간", "거래시간"],
    "kind": ["타입", "구분", "거래구분", "유형"],
    "major": ["대분류", "카테고리"],
    "minor": ["소분류", "세부카테고리"],
    "merchant": ["내용", "거래처", "가맹점", "적요", "거래처명"],
    "amount": ["금액", "거래금액"],
    "currency": ["화폐", "통화"],
    "method": ["결제수단", "자산", "계좌", "카드"],
    "memo": ["메모", "비고"],
}

CATEGORY_EMOJI = {
    "식비": "🍚",
    "카페": "☕",
    "카페/간식": "☕",
    "간식": "🍪",
    "술": "🍺",
    "유흥": "🍺",
    "술/유흥": "🍺",
    "생활": "🧺",
    "생활용품": "🧺",
    "온라인쇼핑": "📦",
    "쇼핑": "🛍️",
    "패션": "👕",
    "패션/쇼핑": "👕",
    "뷰티": "💄",
    "뷰티/미용": "💄",
    "미용": "💄",
    "교통": "🚌",
    "자동차": "🚗",
    "주거": "🏠",
    "통신": "📱",
    "주거/통신": "🏠",
    "의료": "🏥",
    "건강": "🏥",
    "의료/건강": "🏥",
    "금융": "🏦",
    "보험": "🛡️",
    "문화": "🎬",
    "여가": "🎬",
    "문화/여가": "🎬",
    "여행": "✈️",
    "숙박": "🏨",
    "여행/숙박": "✈️",
    "교육": "📚",
    "육아": "🍼",
    "자녀": "🍼",
    "자녀/육아": "🍼",
    "반려동물": "🐾",
    "경조": "🎁",
    "선물": "🎁",
    "경조/선물": "🎁",
    "기부": "💝",
    "세금": "🧾",
    "월세": "🏠",
    "관리비": "🏠",
    "구독": "🔁",
    "급여": "💰",
    "월급": "💰",
    "용돈": "💵",
    "상여": "💰",
    "사업": "💼",
    "사업수입": "💼",
    "금융수입": "📈",
    "이자": "📈",
    "환급": "↩️",
}

DEFAULT_EMOJI_EXPENSE = "💳"
DEFAULT_EMOJI_INCOME = "💰"


@dataclass
class Txn:
    day: date
    at: Optional[time]
    kind: str  # "expense" | "income" | "transfer"
    major: str
    minor: str
    merchant: str
    amount: int
    method: str
    memo: str

    @property
    def emoji(self) -> str:
        for key in (self.major, self.minor):
            if not key:
                continue
            if key in CATEGORY_EMOJI:
                return CATEGORY_EMOJI[key]
            for token in re.split(r"[/,·\s]+", key):
                if token in CATEGORY_EMOJI:
                    return CATEGORY_EMOJI[token]
        return DEFAULT_EMOJI_INCOME if self.kind == "income" else DEFAULT_EMOJI_EXPENSE

    @property
    def title(self) -> str:
        name = self.merchant or self.minor or self.major or "거래"
        return f"{self.emoji} {self.amount:,} {name}".strip()

    @property
    def description(self) -> str:
        lines = []
        category = " › ".join([p for p in (self.major, self.minor) if p])
        if category:
            lines.append(f"분류: {category}")
        if self.method:
            lines.append(f"수단: {self.method}")
        if self.memo:
            lines.append(f"메모: {self.memo}")
        lines.append(f"금액: {self.amount:,}원")
        return "\n".join(lines)

    @property
    def event_id(self) -> str:
        raw = "|".join([
            self.day.isoformat(),
            self.at.strftime("%H:%M") if self.at else "",
            self.kind,
            self.merchant,
            str(self.amount),
            self.method,
        ])
        digest = hashlib.sha256(raw.encode("utf-8")).hexdigest()
        # Google event IDs allow base32hex chars (0-9, a-v); hex digits qualify.
        return "bs" + digest[:40]


XLSX_MARKERS = ("[Content_Types].xml", "xl/workbook.xml")
SHEET_EXTS = (".xlsx", ".xlsm", ".xls")


def _is_xlsx_container(names: list[str]) -> bool:
    """xlsx 자체도 PK로 시작하는 zip이므로, 내부 구조로 구분합니다."""
    return any(marker in names for marker in XLSX_MARKERS)


def _unwrap_zip(raw: bytes, password: str, depth: int) -> bytes:
    with pyzipper.AESZipFile(io.BytesIO(raw)) as zf:
        names = zf.namelist()

        if _is_xlsx_container(names):
            return raw  # 이미 엑셀 파일

        candidates = [
            n for n in names
            if n.lower().endswith(SHEET_EXTS) and not n.startswith("__MACOSX")
        ]
        if not candidates:
            raise ValueError(
                f"zip 안에 엑셀 파일이 없습니다. 내용물: {names[:10]}"
            )
        if len(candidates) > 1:
            log.warning("zip 안에 엑셀이 %d개, 첫 번째만 사용합니다.", len(candidates))
        target = candidates[0]
        log.info("zip에서 추출: %s", target)

        pwd = password.encode("utf-8") if password else None
        try:
            inner = zf.read(target, pwd=pwd)
        except RuntimeError as err:
            if "password" in str(err).lower():
                raise ValueError(
                    "zip 비밀번호가 틀렸거나 EXCEL_PASSWORD가 비어 있습니다."
                ) from err
            raise

    return unwrap(inner, password, depth + 1)


def unwrap(raw: bytes, password: str = "", depth: int = 0) -> bytes:
    """zip 포장을 벗기고 엑셀 암호를 풀어 순수 xlsx 바이트를 반환합니다.

    뱅크샐러드는 비밀번호가 걸린 zip 안에 엑셀을 담아 보냅니다.
    포장 없이 xlsx만 오는 경우도 그대로 처리합니다.
    """
    if depth > 3:
        raise ValueError("압축이 너무 여러 겹입니다.")

    if raw[:2] == b"PK":
        return _unwrap_zip(raw, password, depth)

    # OLE/CFB 헤더 = 비밀번호가 걸린 오피스 파일
    if not password:
        raise ValueError("암호화된 엑셀인데 EXCEL_PASSWORD가 비어 있습니다.")
    out = io.BytesIO()
    office = msoffcrypto.OfficeFile(io.BytesIO(raw))
    try:
        office.load_key(password=password)
        office.decrypt(out)
    except Exception as err:
        raise ValueError(f"엑셀 비밀번호 해제 실패: {err}") from err
    return unwrap(out.getvalue(), password, depth + 1)


# 하위 호환용 별칭
decrypt_if_needed = unwrap


def _pick_sheet(wb) -> "openpyxl.worksheet.worksheet.Worksheet":
    names = wb.sheetnames
    for hint in SHEET_HINTS:
        for name in names:
            if hint in name.replace(" ", ""):
                return wb[name]
    # Fall back to the sheet with the most rows.
    return max((wb[n] for n in names), key=lambda ws: ws.max_row)


def _norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _find_header(rows) -> tuple[int, dict[str, int]]:
    for idx, row in enumerate(rows[:15]):
        cells = [_norm(c).replace(" ", "") for c in row]
        if not any(c in ("날짜", "거래일자", "일자", "거래일") for c in cells):
            continue
        if not any("금액" in c for c in cells):
            continue
        mapping: dict[str, int] = {}
        for field, aliases in COLUMN_ALIASES.items():
            for col, cell in enumerate(cells):
                if cell in aliases and field not in mapping:
                    mapping[field] = col
        return idx, mapping
    raise ValueError("가계부 내역 시트에서 헤더 행(날짜·금액)을 찾지 못했습니다.")


def _to_date(value) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _norm(value)
    if not text:
        return None
    text = text.split(" ")[0]
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y년%m월%d일", "%Y%m%d"):
        try:
            return datetime.strptime(text.replace(" ", ""), fmt).date()
        except ValueError:
            continue
    return None


def _to_time(value) -> Optional[time]:
    if isinstance(value, datetime):
        parsed = value.time()
    elif isinstance(value, time):
        parsed = value
    else:
        text = _norm(value)
        if not text:
            return None
        match = re.search(r"(\d{1,2}):(\d{2})(?::(\d{2}))?", text)
        if match:
            parsed = time(int(match.group(1)), int(match.group(2)))
        elif re.fullmatch(r"\d{4}", text):
            parsed = time(int(text[:2]), int(text[2:]))
        else:
            return None
    if parsed.hour == 0 and parsed.minute == 0:
        return None
    return parsed.replace(second=0, microsecond=0)


def _to_amount(value) -> Optional[int]:
    if isinstance(value, (int, float)):
        return int(round(abs(value)))
    text = _norm(value)
    if not text:
        return None
    text = re.sub(r"[^\d.\-]", "", text)
    if not text or text in ("-", "."):
        return None
    try:
        return int(round(abs(float(text))))
    except ValueError:
        return None


def _classify(kind_text: str, amount_raw) -> str:
    text = kind_text.replace(" ", "")
    if "이체" in text or "송금" in text:
        return "transfer"
    if "수입" in text or "입금" in text:
        return "income"
    if "지출" in text or "출금" in text:
        return "expense"
    if isinstance(amount_raw, (int, float)):
        return "income" if amount_raw > 0 else "expense"
    return "expense"


def parse(raw: bytes, password: str = "") -> Iterator[Txn]:
    data = unwrap(raw, password)
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    try:
        ws = _pick_sheet(wb)
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    header_idx, cols = _find_header(rows)
    log.info("헤더 %d행, 매핑된 컬럼: %s", header_idx + 1, sorted(cols))

    if "date" not in cols or "amount" not in cols:
        raise ValueError("날짜 또는 금액 컬럼을 찾지 못했습니다.")

    def cell(row, field):
        idx = cols.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    for row in rows[header_idx + 1:]:
        if row is None or all(c is None or _norm(c) == "" for c in row):
            continue
        day = _to_date(cell(row, "date"))
        if day is None:
            continue
        amount_raw = cell(row, "amount")
        amount = _to_amount(amount_raw)
        if amount is None:
            continue
        yield Txn(
            day=day,
            at=_to_time(cell(row, "time")),
            kind=_classify(_norm(cell(row, "kind")), amount_raw),
            major=_norm(cell(row, "major")),
            minor=_norm(cell(row, "minor")),
            merchant=_norm(cell(row, "merchant")),
            amount=amount,
            method=_norm(cell(row, "method")),
            memo=_norm(cell(row, "memo")),
        )
